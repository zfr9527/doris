#!/usr/bin/env python3
import re
import openpyxl

LOG_FILE = '/Users/zhangfurong/Documents/work/my_repo/doris/regression-test/suites/nereids_p0/shuffle_key_prune/logs/doris-regression-test.20260320.143747.log'
XLSX_FILE = '/Users/zhangfurong/Documents/work/my_repo/doris/regression-test/suites/nereids_p0/shuffle_key_prune/agg_shuffle_key_data.xlsx'

# Parse log
data = {}  # label -> {times_close, times_open, eqClose, eqOpen}

with open(LOG_FILE, 'r') as f:
    for line in f:
        m = re.search(r'(\w+_\w+_\w+_\w+) times_close=\[(.+?)\]', line)
        if m:
            label = m.group(1)
            vals = [int(x.strip()) for x in m.group(2).split(',')]
            data.setdefault(label, {})['times_close'] = vals
            continue
        m = re.search(r'(\w+_\w+_\w+_\w+) times_open=\[(.+?)\]', line)
        if m:
            label = m.group(1)
            vals = [int(x.strip()) for x in m.group(2).split(',')]
            data.setdefault(label, {})['times_open'] = vals
            continue
        m = re.search(r'(\w+_\w+_\w+_\w+) equivalenceExprIdsClose=(.+)', line)
        if m:
            label = m.group(1)
            raw = m.group(2).strip()
            data.setdefault(label, {})['eqClose'] = raw
            continue
        m = re.search(r'(\w+_\w+_\w+_\w+) equivalenceExprIdsOpen=(.+)', line)
        if m:
            label = m.group(1)
            raw = m.group(2).strip()
            data.setdefault(label, {})['eqOpen'] = raw
            continue

def parse_eq_list(raw_str):
    """Parse the outer list, splitting by sql index (14 items)"""
    items = []
    depth = 0
    current = ''
    for ch in raw_str:
        if ch == '[' and depth == 0:
            depth += 1
            continue
        elif ch == '[':
            depth += 1
            current += ch
        elif ch == ']' and depth == 1:
            items.append(current.strip())
            current = ''
            depth -= 1
        elif ch == ']':
            depth -= 1
            current += ch
        elif ch == ',' and depth == 1:
            continue
        else:
            current += ch
    return items

gby_ids = ['gby2', 'gby3', 'gby4', 'gby5', 'gby6']
table_ids = ['dist_ndv_low_tb', 'dist_ndv_high_tb', 'random_ndv_low_tb', 'random_ndv_high_tb']

wb = openpyxl.load_workbook(XLSX_FILE)

# Sheet 1: 耗时数据
sheet_name_time = 'multi_gby_new_耗时数据'
if sheet_name_time in wb.sheetnames:
    del wb[sheet_name_time]
ws_time = wb.create_sheet(sheet_name_time)

header = ['sql_id', 'table_id']
for g in gby_ids:
    n = g.replace('gby', '')
    header.extend([f'on_{n}', f'off_{n}'])
ws_time.append(header)

for tb in table_ids:
    for sql_idx in range(14):
        row = [sql_idx, tb]
        for gby in gby_ids:
            label = f'{gby}_{tb}'
            if label in data and 'times_close' in data[label] and 'times_open' in data[label]:
                tc = data[label]['times_close']
                to = data[label]['times_open']
                if sql_idx < len(tc):
                    row.extend([tc[sql_idx], to[sql_idx]])
                else:
                    row.extend(['', ''])
            else:
                row.extend(['', ''])
        ws_time.append(row)

# Sheet 2: 等价类数据
sheet_name_eq = 'multi_gby_new_等价类数据'
if sheet_name_eq in wb.sheetnames:
    del wb[sheet_name_eq]
ws_eq = wb.create_sheet(sheet_name_eq)
ws_eq.append(header)

for tb in table_ids:
    for sql_idx in range(14):
        row = [sql_idx, tb]
        for gby in gby_ids:
            label = f'{gby}_{tb}'
            if label in data:
                eq_close_raw = data[label].get('eqClose', '[]')
                eq_open_raw = data[label].get('eqOpen', '[]')
                eq_close_items = parse_eq_list(eq_close_raw)
                eq_open_items = parse_eq_list(eq_open_raw)
                ec = eq_close_items[sql_idx] if sql_idx < len(eq_close_items) else ''
                eo = eq_open_items[sql_idx] if sql_idx < len(eq_open_items) else ''
                row.extend([ec, eo])
            else:
                row.extend(['', ''])
        ws_eq.append(row)

wb.save(XLSX_FILE)
print(f'Done! Added sheets: {sheet_name_time}, {sheet_name_eq}')
print(f'Labels found: {sorted(data.keys())}')
print(f'Total labels: {len(data)}')
for label in sorted(data.keys()):
    d = data[label]
    tc = d.get('times_close', [])
    to = d.get('times_open', [])
    print(f'  {label}: close={tc}, open={to}')
